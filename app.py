from flask import Flask, render_template, request, jsonify, send_file
import json
import os
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)

DATA_FILE = 'data.json'

def load_data():
    if not os.path.exists(DATA_FILE):
        notas = [
            {"id": 1, "numero": "1045", "fornecedor": "Construtora ABC", "valor": 15500, "status": "processando", "observacao": "Status", "dias": 2},
            {"id": 2, "numero": "1049", "fornecedor": "Ang. Obras",       "valor": 19500, "status": "processando", "observacao": "Conferindo", "dias": 1},
            {"id": 3, "numero": "1048", "fornecedor": "Material Forte",   "valor": 8900,  "status": "medicao",     "observacao": "Conferindo", "dias": 1},
            {"id": 4, "numero": "1042", "fornecedor": "Logística Sol",    "valor": 22150, "status": "pendencia",   "observacao": "Divergência Valor", "dias": 5},
            {"id": 5, "numero": "1051", "fornecedor": "Eng. Tech",        "valor": 11200, "status": "pendencia",   "observacao": "Falta Medição", "dias": 3},
            {"id": 6, "numero": "1039", "fornecedor": "Tech Obras",       "valor": 29800, "status": "concluida",   "observacao": "Conferido", "dias": 0},
            {"id": 7, "numero": "1055", "fornecedor": "Elétrica J.J.",    "valor": 7600,  "status": "concluida",   "observacao": "Lançado", "dias": 0},
        ]
        save_data(notas)
        return notas
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(notas):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(notas, f, ensure_ascii=False, indent=2)

def next_id(notas):
    return max((n['id'] for n in notas), default=0) + 1

# ── Rotas ──────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/notas', methods=['GET'])
def get_notas():
    return jsonify(load_data())

@app.route('/api/notas', methods=['POST'])
def create_nota():
    notas = load_data()
    body = request.get_json()
    nova = {
        "id": next_id(notas),
        "numero": body.get('numero', ''),
        "fornecedor": body.get('fornecedor', ''),
        "valor": float(body.get('valor', 0)),
        "status": body.get('status', 'processando'),
        "observacao": body.get('observacao', ''),
        "dias": int(body.get('dias', 0)),
    }
    notas.append(nova)
    save_data(notas)
    return jsonify(nova), 201

@app.route('/api/notas/<int:nid>/status', methods=['PATCH'])
def update_status(nid):
    notas = load_data()
    body = request.get_json()
    for n in notas:
        if n['id'] == nid:
            n['status'] = body.get('status', n['status'])
            save_data(notas)
            return jsonify(n)
    return jsonify({"erro": "não encontrado"}), 404

@app.route('/api/notas/<int:nid>', methods=['DELETE'])
def delete_nota(nid):
    notas = load_data()
    notas = [n for n in notas if n['id'] != nid]
    save_data(notas)
    return jsonify({"ok": True})

@app.route('/api/exportar', methods=['GET'])
def exportar_excel():
    notas = load_data()

    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    # Cores
    fill_verde   = PatternFill("solid", fgColor="C6EFCE")  # concluida
    fill_laranja = PatternFill("solid", fgColor="FFEB9C")  # processando / medicao
    fill_vermelho = PatternFill("solid", fgColor="FFC7CE") # pendencia
    fill_header  = PatternFill("solid", fgColor="1B2232")

    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_normal = Font(size=10)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers = ["NF", "Fornecedor", "Valor (R$)", "Status", "Observação", "Dias"]
    col_widths = [10, 28, 16, 16, 28, 8]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    status_label = {
        'processando': 'Em Processamento',
        'medicao': 'Em Medição',
        'pendencia': 'Com Pendência',
        'concluida': 'Concluída'
    }

    for row_idx, n in enumerate(notas, 2):
        valores = [
            f"#{n['numero']}",
            n['fornecedor'],
            n['valor'],
            status_label.get(n['status'], n['status']),
            n.get('observacao', ''),
            n.get('dias', 0)
        ]
        if n['status'] == 'concluida':
            fill = fill_verde
        elif n['status'] == 'pendencia':
            fill = fill_vermelho
        else:
            fill = fill_laranja

        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = font_normal
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='notas_fiscais.xlsx'
    )


@app.route('/api/importar', methods=['POST'])
def importar_excel():
    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files['arquivo']
    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"erro": "Formato inválido. Use .xlsx ou .xls"}), 400

    try:
        wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
        ws = wb.active

        notas_existentes = load_data()
        ids_existentes = {n['numero']: n for n in notas_existentes}

        status_map = {
            'em processamento': 'processando',
            'processando': 'processando',
            'em medição': 'medicao',
            'medição': 'medicao',
            'medicao': 'medicao',
            'com pendência': 'pendencia',
            'pendência': 'pendencia',
            'pendencia': 'pendencia',
            'concluída': 'concluida',
            'concluida': 'concluida',
            'lançado': 'concluida',
        }

        importadas = 0
        atualizadas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Ignora linhas vazias
            if not any(row):
                continue

            # Tenta extrair campos pela posição das colunas da planilha enviada
            # Colunas: NF, Fornecedor, Valor, Status, Observação, Dias
            try:
                numero     = str(row[0]).strip().replace('#', '') if row[0] else None
                fornecedor = str(row[1]).strip() if row[1] else ''
                valor      = float(str(row[2]).replace('R$','').replace('.','').replace(',','.').strip()) if row[2] else 0
                status_raw = str(row[3]).strip().lower() if row[3] else 'processando'
                observacao = str(row[4]).strip() if row[4] else ''
                dias       = int(row[5]) if row[5] else 0
            except Exception:
                continue

            if not numero or numero == 'None':
                continue

            status = status_map.get(status_raw, 'processando')

            if numero in ids_existentes:
                # Atualiza nota existente
                n = ids_existentes[numero]
                n['fornecedor'] = fornecedor
                n['valor']      = valor
                n['status']     = status
                n['observacao'] = observacao
                n['dias']       = dias
                atualizadas += 1
            else:
                # Cria nova nota
                nova = {
                    "id": next_id(notas_existentes),
                    "numero": numero,
                    "fornecedor": fornecedor,
                    "valor": valor,
                    "status": status,
                    "observacao": observacao,
                    "dias": dias,
                }
                notas_existentes.append(nova)
                ids_existentes[numero] = nova
                importadas += 1

        save_data(notas_existentes)
        return jsonify({
            "ok": True,
            "importadas": importadas,
            "atualizadas": atualizadas,
            "total": len(notas_existentes)
        })

    except Exception as e:
        return jsonify({"erro": str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
